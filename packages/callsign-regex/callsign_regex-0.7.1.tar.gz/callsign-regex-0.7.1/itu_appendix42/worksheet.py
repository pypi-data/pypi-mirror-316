""" itu_appendix42 worksheet.py
Copyright (C) 2024 Martin J Levy - W6LHI/G8LHI - @mahtin - https://github.com/mahtin
"""

import sys
from os import access, R_OK
from os.path import isfile, join, expanduser, getmtime
from glob import glob

from zipfile import BadZipFile
from openpyxl import load_workbook

try:
    from importlib.resources import files
except ImportError:
    from importlib_resources import files

class ItuAppendix42WorksheetError(Exception):
    """ ItuAppendix42 errors """

class ItuAppendix42Worksheet():
    """ ItuAppendix42 Worksheet """

    DOWNLOAD_FOLDER = 'Downloads'
    PACKAGE_RESOURCES = 'itu_appendix42.resources'

    FILENAME_PATTERN = 'CallSignSeriesRanges-*-*-*-*.xlsx'

    WORKSHEET_NAME = 'Exported data'

    _log = None

    def __init__(self, log=None):
        """ ItuAppendix42 Worksheet """
        self.__class__._log = log

    def read_sheet(self):
        """ read_sheet """
        worksheet = self._find_worksheet_and_read_in()
        # skip first line
        v =  list(worksheet.values)[1:]
        self.__class__._log.info('spreadsheet values [%s ...]', v[0])
        return v

    def _package_folder(self):
        """ _package_folder """
        return str(files(self.__class__.PACKAGE_RESOURCES)._paths[0])

    def _download_folder(self):
        """ _download_folder """
        return str(join(expanduser('~'), self.__class__.DOWNLOAD_FOLDER))

    def _find_newest_filename(self):
        """ _find_filename """

        #
        # Change in version 3.10 becuase of the added the root_dir and dir_fd parameters ...
        # filenames = glob(cls.FILENAME_PATTERN, root_dir=cls._package_folder()) + glob(cls.FILENAME_PATTERN, root_dir=dirname)
        # ... but we leave it in the old format so it works on earlier Python versions.
        #

        all_filenames = []
        for filename in glob(self._package_folder() + '/' + self.__class__.FILENAME_PATTERN) + glob(self._download_folder() + '/' + self.__class__.FILENAME_PATTERN):
            try:
                mtime = getmtime(filename)
            except OSError as e:
                # we simply ignore this file
                self.__class__._log.info('ignore %s', filename)
                continue
            if not isfile(filename) or not access(filename, R_OK):
                self.__class__._log.info('ignore %s (not readable)', filename)
                continue
            all_filenames.append((filename, mtime))
            self.__class__._log.info('%d %s', mtime, filename)
        all_filenames = sorted(all_filenames, key=lambda item: item[1])
        if len(all_filenames) == 0:
            raise ItuAppendix42WorksheetError(self.__class__.FILENAME_PATTERN + ': not found') from None
        best = all_filenames[-1]
        filename = best[0]
        self.__class__._log.info('using %s', filename)
        return filename

    def _find_worksheet_and_read_in(self):
        """ _find_worksheet_and_read_in """
        try:
            filename = self._find_newest_filename()
        except ItuAppendix42WorksheetError as e:
            self.__class__._log.warning('%s', e)
            raise ItuAppendix42WorksheetError(self.__class__.FILENAME_PATTERN) from None

        try:
            workbook = load_workbook(filename=filename, data_only=True)
        except OSError as e:
            self.__class__._log.warning('%s: %s', filename, e)
            raise ItuAppendix42WorksheetError(filename)
        except BadZipFile as e:
            self.__class__._log.warning('%s: %s', filename, e)
            raise ItuAppendix42WorksheetError(filename)

        if self.__class__.WORKSHEET_NAME != workbook.sheetnames[0]:
            self.__class__._log.warning('%s: missing worksheet %s', filename, self.__class__.WORKSHEET_NAME)
            raise ItuAppendix42WorksheetError(filename)
        #worksheet = workbook.active
        worksheet = workbook[self.__class__.WORKSHEET_NAME]
        return worksheet

