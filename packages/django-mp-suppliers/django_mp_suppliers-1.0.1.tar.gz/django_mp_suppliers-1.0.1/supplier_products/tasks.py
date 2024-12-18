
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

from openpyxl import load_workbook

from manufacturers.utils import ManufacturerCollection

from supplier_products.models import SupplierProduct, ImportTask

from exchange.models import ExchangeRates

from core import celery_app


def _get_percent(i, total, max_val=100, decimals=1, ):
    return ("{0:." + str(decimals) + "f}").format(max_val * (i / float(total)))


@celery_app.task(bind=True)
def import_supplier_products(self, task_id, col_map):

    task = ImportTask.objects.get(pk=task_id)

    task.update_progress(1, _('Loading file'), is_processing=True)

    worksheet = load_workbook(task.file.path).worksheets[0]

    task.update_progress(15, _('Loading manufacturers'))

    manufacturers = ManufacturerCollection()

    rates = ExchangeRates.objects.get()

    if task.should_data_be_cleaned:
        task.supplier.clean_products()

    task.update_progress(20, _('Loading products'))

    row_count = worksheet.max_row

    for i, value in enumerate(worksheet.values):

        values = {
            'index': str(value[col_map['index']]),
            'price': str(value[col_map['price']]),
            'stock': str(value[col_map['stock']]),
            'supplier': task.supplier,
            'warehouse': task.warehouse,
        }

        if col_map.get('manufacturer'):
            values['manufacturer_id'] = manufacturers.get(
                value[col_map['manufacturer']])

        if col_map.get('description'):
            values['description'] = value[col_map['description']]

        try:
            _import_product(
                check_existing=not task.should_data_be_cleaned,
                rates=rates,
                **values)
        except Exception as e:
            print(e)

        if i % 10 == 0:
            task.update_progress(20 + float(_get_percent(i, row_count, 70)))

    task.update_progress(
        percent=100,
        status=_('Done'),
        is_processing=False,
        is_completed=True
    )

    task.supplier.price_updated = timezone.now()
    task.supplier.save(update_fields=['price_updated'])


def _import_product(
        supplier,
        warehouse,
        index,
        price,
        stock,
        manufacturer_id=None,
        description='',
        check_existing=True,
        rates=None):

    assert rates is not None

    if not stock or stock == '0' or not price or not index:
        return

    product = None
    index = index.strip()

    if check_existing and index:
        try:
            product = SupplierProduct.objects.get(
                index=index,
                supplier=supplier,
                warehouse=warehouse)
        except SupplierProduct.DoesNotExist:
            pass

    if product is None:
        product = SupplierProduct(
            index=index,
            supplier=supplier,
            warehouse=warehouse)

    price = float(price.replace(',', '.'))

    product.price_retail = price
    product.price_wholesale = price
    product.currency_initial = supplier.currency
    product.stock = stock

    if manufacturer_id:
        product.manufacturer_id = manufacturer_id

    if description:
        product.description = description

    product.save()
